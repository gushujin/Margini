"""
Calcolatrice Commerciale — app.py
Tutti i colori, font, dimensioni e testi sono letti da config.xlsx
Non modificare valori visivi o label direttamente in questo file.
"""

import streamlit as st
import openpyxl

st.set_page_config(
    page_title="Calcolatrice Commerciale",
    page_icon="🧮",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ══════════════════════════════════════════════════════════════════════════════
# CARICAMENTO CONFIGURAZIONE DA config.xlsx
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data
def load_config():
    wb = openpyxl.load_workbook("config.xlsx", data_only=True)

    colori = {}
    for row in wb["COLORI"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            colori[str(row[0]).strip()] = str(row[1]).strip()

    font = {}
    for row in wb["FONT"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            font[str(row[0]).strip()] = str(row[1]).strip()

    testi = {}
    for row in wb["TESTI"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            testi[str(row[0]).strip()] = row[1]

    return colori, font, testi


C, F, T = load_config()

# ══════════════════════════════════════════════════════════════════════════════
# CSS GENERATO DINAMICAMENTE DA CONFIGURAZIONE
# ══════════════════════════════════════════════════════════════════════════════

weights_display = F["font_display_weights"].replace(";", ";")
weights_mono    = F["font_mono_weights"].replace(";", ";")
google_url = (
    f"https://fonts.googleapis.com/css2?"
    f"family={F['font_display'].replace(' ', '+')}:wght@{weights_display}"
    f"&family={F['font_mono'].replace(' ', '+')}:wght@{weights_mono}"
    f"&display=swap"
)

css = f"""
<style>
@import url('{google_url}');

:root {{
    --bg:        {C['bg']};
    --surface:   {C['surface']};
    --border:    {C['border']};
    --accent:    {C['accent']};
    --accent2:   {C['accent2']};
    --text:      {C['text']};
    --muted:     {C['muted']};
    --danger:    {C['danger']};
    --ok:        {C['ok']};
    --f-display: '{F['font_display']}', sans-serif;
    --f-mono:    '{F['font_mono']}', monospace;
    --sz-title:      {F['size_title']};
    --sz-section:    {F['size_section']};
    --sz-result:     {F['size_result']};
    --sz-result-sec: {F['size_result_sec']};
    --sz-label:      {F['size_label']};
    --sz-mono-sm:    {F['size_mono_small']};
}}

html, body, [data-testid="stAppViewContainer"] {{
    background-color: var(--bg) !important;
    color: var(--text);
    font-family: var(--f-display);
}}

[data-testid="stHeader"]  {{ background: transparent !important; }}
[data-testid="stSidebar"] {{ background: var(--surface) !important; }}
#MainMenu, footer, header {{ visibility: hidden; }}

/* ── Top bar ── */
.topbar {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 1.5rem 0 2rem;
    border-bottom: 1px solid var(--border);
    margin-bottom: 2rem;
}}
.topbar h1 {{
    font-family: var(--f-display);
    font-size: var(--sz-title);
    font-weight: 800;
    letter-spacing: -0.03em;
    margin: 0;
    color: var(--text);
}}
.topbar .tag {{
    background: var(--accent);
    color: #000;
    font-family: var(--f-mono);
    font-size: var(--sz-mono-sm);
    font-weight: 500;
    padding: 0.3rem 0.7rem;
    border-radius: 2px;
    letter-spacing: 0.1em;
    text-transform: uppercase;
}}

/* ── Section labels ── */
.section-label {{
    font-family: var(--f-mono);
    font-size: var(--sz-mono-sm);
    letter-spacing: 0.18em;
    text-transform: uppercase;
    color: var(--accent);
    margin-bottom: 0.4rem;
}}
.section-title {{
    font-size: var(--sz-section);
    font-weight: 700;
    margin-bottom: 1.25rem;
    color: var(--text);
}}

/* ── Result boxes ── */
.result-box {{
    background: #111;
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 1rem 1.2rem;
    margin-top: 0.5rem;
    margin-bottom: 0.25rem;
}}
.result-label {{
    font-family: var(--f-mono);
    font-size: var(--sz-mono-sm);
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 0.25rem;
}}
.result-value          {{ font-size: var(--sz-result);     font-weight: 800; color: var(--accent);  letter-spacing: -0.02em; font-family: var(--f-display); }}
.result-value.secondary{{ font-size: var(--sz-result-sec); color: var(--text); }}
.result-value.ok-val   {{ color: var(--ok); }}
.result-value.warn-val {{ color: var(--accent2); }}
.result-value.danger-v {{ color: var(--danger); }}

/* ── Alert box ── */
.alert-box {{
    background: rgba(255,59,59,0.08);
    border: 1px solid rgba(255,59,59,0.35);
    border-left: 3px solid var(--danger);
    border-radius: 6px;
    padding: 1rem 1.2rem;
    margin-top: 1rem;
}}
.alert-title {{
    color: var(--danger);
    font-weight: 700;
    font-size: 0.9rem;
    margin-bottom: 0.4rem;
    font-family: var(--f-display);
}}
.alert-body {{
    color: var(--text);
    font-size: 0.82rem;
    line-height: 1.7;
    font-family: var(--f-mono);
}}

/* ── Discount preview chip ── */
.disc-chip {{
    font-family: var(--f-mono);
    font-size: 0.72rem;
    color: var(--accent);
    background: rgba(212,255,0,0.07);
    border: 1px solid rgba(212,255,0,0.2);
    border-radius: 4px;
    padding: 0.25rem 0.6rem;
    margin-top: 0.2rem;
    display: inline-block;
}}

/* ── Streamlit input overrides ── */
div[data-testid="stNumberInput"] input,
div[data-testid="stTextInput"] input {{
    background: #111 !important;
    border: 1px solid var(--border) !important;
    color: var(--text) !important;
    border-radius: 4px !important;
    font-family: var(--f-mono) !important;
    font-size: 0.95rem !important;
}}
div[data-testid="stNumberInput"] input:focus,
div[data-testid="stTextInput"] input:focus {{
    border-color: var(--accent) !important;
    box-shadow: 0 0 0 2px rgba(212,255,0,0.15) !important;
}}
label[data-testid="stWidgetLabel"] {{
    color: var(--muted) !important;
    font-family: var(--f-mono) !important;
    font-size: var(--sz-label) !important;
    letter-spacing: 0.05em !important;
    text-transform: uppercase !important;
}}

/* ── Tabs ── */
button[data-baseweb="tab"] {{
    font-family: var(--f-display) !important;
    font-weight: 600 !important;
    font-size: 0.85rem !important;
    color: var(--muted) !important;
    background: transparent !important;
    border-bottom: 2px solid transparent !important;
}}
button[data-baseweb="tab"][aria-selected="true"] {{
    color: var(--accent) !important;
    border-bottom-color: var(--accent) !important;
}}
[data-testid="stTabs"] [role="tablist"] {{
    border-bottom: 1px solid var(--border) !important;
    gap: 0.5rem !important;
}}

/* ── Reset button ── */
div[data-testid="stButton"] > button {{
    background: transparent !important;
    border: 1px solid var(--border) !important;
    color: var(--muted) !important;
    font-family: var(--f-mono) !important;
    font-size: 0.75rem !important;
    letter-spacing: 0.08em !important;
    border-radius: 4px !important;
    transition: all 0.15s ease !important;
}}
div[data-testid="stButton"] > button:hover {{
    border-color: var(--accent) !important;
    color: var(--accent) !important;
    background: rgba(212,255,0,0.05) !important;
}}

hr {{ border-color: var(--border) !important; }}
[data-testid="column"] {{ padding: 0 0.5rem !important; }}
</style>
"""
st.markdown(css, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def parse_sconto(s: str):
    """Parsa uno sconto semplice o concatenato '40+10+5' → float 0-1 oppure None."""
    if not s or not s.strip():
        return None
    try:
        netto = 100.0
        for parte in s.strip().split("+"):
            v = float(parte.strip())
            if v < 0 or v >= 100:
                return None
            netto *= (1 - v / 100)
        return 1 - netto / 100
    except (ValueError, ZeroDivisionError):
        return None


def fmt_pct(v: float) -> str:
    return f"{v * 100:.2f}%"


def fmt_eur(v: float) -> str:
    return f"€ {v:,.2f}"


def chip_sconto(raw: str):
    """Mostra il chip con sconto totale se la stringa contiene '+'."""
    if raw and "+" in raw:
        v = parse_sconto(raw)
        if v is not None:
            st.markdown(
                f'<div class="disc-chip">→ Sconto totale: {v*100:.2f}%</div>',
                unsafe_allow_html=True,
            )


def result_box(label: str, value: str, cls: str = ""):
    st.markdown(
        f'<div class="result-box">'
        f'<div class="result-label">{label}</div>'
        f'<div class="result-value {cls}">{value}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )


def alert_margine(acquisto: float, vendita: float):
    """
    Mostra alert se margine < soglia e calcola la riduzione necessaria.
    Soglia letta da config: T['soglia_margine'] (numero intero es. 10).
    """
    soglia = float(T["soglia_margine"]) / 100
    if vendita <= 0:
        return
    margine = (vendita - acquisto) / vendita
    if margine < soglia:
        acq_max   = vendita * (1 - soglia)
        riduzione = acquisto - acq_max
        st.markdown(
            f'<div class="alert-box">'
            f'<div class="alert-title">{T["alert_title"]}</div>'
            f'<div class="alert-body">'
            f'Margine attuale: <strong>{margine*100:.2f}%</strong> '
            f'(soglia: {T["soglia_margine"]}%)<br>'
            f'{T["alert_acq_max"]}: <strong>{fmt_eur(acq_max)}</strong><br>'
            f'{T["alert_riduzione"]}: <strong>{fmt_eur(riduzione)}</strong>'
            f'</div></div>',
            unsafe_allow_html=True,
        )


def alert_margine_da_sconti(sc_acq: float, sc_ven: float):
    """Alert margine < soglia calcolato da sconti %."""
    soglia = float(T["soglia_margine"]) / 100
    if (1 - sc_ven) == 0:
        return
    margine = (sc_acq - sc_ven) / (1 - sc_ven)
    if margine < soglia:
        sc_acq_max = 1 - (1 - sc_ven) / (1 - soglia)
        st.markdown(
            f'<div class="alert-box">'
            f'<div class="alert-title">{T["alert_title"]}</div>'
            f'<div class="alert-body">'
            f'Margine attuale: <strong>{margine*100:.2f}%</strong> '
            f'(soglia: {T["soglia_margine"]}%)<br>'
            f'Sconto acquisto massimo consentito: '
            f'<strong>{sc_acq_max*100:.2f}%</strong><br>'
            f'(con sconto vendita fisso a {sc_ven*100:.2f}%)'
            f'</div></div>',
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STATE — reset ad ogni avvio, nessun valore preimpostato
# ══════════════════════════════════════════════════════════════════════════════

KEYS_NUM  = ["s1_acq","s1_ven","s2_listino","s3_acq","s3_marg","s4_listino","s4_netto"]
KEYS_TEXT = ["s2_sc_acq","s2_sc_ven","s5_sc"]
ALL_KEYS  = KEYS_NUM + KEYS_TEXT

def reset_all():
    for k in ALL_KEYS:
        if k in st.session_state:
            del st.session_state[k]

# Inizializza alla prima apertura
for k in KEYS_NUM:
    if k not in st.session_state:
        st.session_state[k] = None
for k in KEYS_TEXT:
    if k not in st.session_state:
        st.session_state[k] = ""


# ══════════════════════════════════════════════════════════════════════════════
# TOP BAR
# ══════════════════════════════════════════════════════════════════════════════

col_h, col_btn = st.columns([5, 1])
with col_h:
    st.markdown(
        f'<div class="topbar">'
        f'<h1>{T["app_title"]}</h1>'
        f'<span class="tag">{T["app_version"]}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )
with col_btn:
    st.markdown("<div style='padding-top:1.9rem'>", unsafe_allow_html=True)
    if st.button(str(T["btn_reset"])):
        reset_all()
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# TABS — nomi da configurazione
# ══════════════════════════════════════════════════════════════════════════════

tab1, tab2, tab3, tab4, tab5 = st.tabs([
    str(T["tab1_nome"]),
    str(T["tab2_nome"]),
    str(T["tab3_nome"]),
    str(T["tab4_nome"]),
    str(T["tab5_nome"]),
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Margine da Prezzi €
# Formule: MARGINE=(Ven-Acq)/Ven  RICARICO=(Ven/Acq)-1
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.markdown(
        f'<div class="section-label">Sezione {T["tab1_label_num"]}</div>'
        f'<div class="section-title">{T["tab1_titolo"]}</div>',
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    with c1:
        acq1 = st.number_input(
            str(T["tab1_input1"]), min_value=0.0, step=0.01, format="%.2f",
            key="s1_acq", value=None, placeholder="es. 50.00",
        )
    with c2:
        ven1 = st.number_input(
            str(T["tab1_input2"]), min_value=0.0, step=0.01, format="%.2f",
            key="s1_ven", value=None, placeholder="es. 60.00",
        )

    if acq1 is not None and ven1 is not None:
        if ven1 == 0:
            st.error("Il prezzo di vendita non può essere zero.")
        elif acq1 > ven1:
            st.warning("⚠ Il prezzo di acquisto è maggiore del prezzo di vendita.")
        else:
            margine  = (ven1 - acq1) / ven1                    # MARGINE
            ricarico = (ven1 / acq1 - 1) if acq1 > 0 else 0   # RICARICO
            utile    = ven1 - acq1

            ca, cb, cc = st.columns(3)
            with ca:
                cls = "ok-val" if margine >= float(T["soglia_margine"]) / 100 else "danger-v"
                result_box(str(T["tab1_res1"]), fmt_pct(margine), cls)
            with cb:
                result_box(str(T["tab1_res2"]), fmt_pct(ricarico), "secondary")
            with cc:
                result_box(str(T["tab1_res3"]), fmt_eur(utile), "secondary")

            alert_margine(acq1, ven1)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Margine da Sconti %
# Formule: MARGINE=((1-Sc_Ven)-(1-Sc_Acq))/(1-Sc_Ven)
#          RICARICO=(1-Sc_Ven)/(1-Sc_Acq)-1
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown(
        f'<div class="section-label">Sezione {T["tab2_label_num"]}</div>'
        f'<div class="section-title">{T["tab2_titolo"]}</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        f"<small style='color:{C['muted']}'>{T['tab2_hint']}</small><br><br>",
        unsafe_allow_html=True,
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        sc_acq_raw = st.text_input(
            str(T["tab2_input1"]), key="s2_sc_acq", placeholder="es. 45  oppure  40+10",
        )
        chip_sconto(sc_acq_raw)
    with c2:
        sc_ven_raw = st.text_input(
            str(T["tab2_input2"]), key="s2_sc_ven", placeholder="es. 30  oppure  25+5",
        )
        chip_sconto(sc_ven_raw)
    with c3:
        listino2 = st.number_input(
            str(T["tab2_input3"]), min_value=0.0, step=0.01, format="%.2f",
            key="s2_listino", value=None, placeholder="facoltativo",
        )

    if sc_acq_raw and sc_ven_raw:
        sc_acq = parse_sconto(sc_acq_raw)
        sc_ven = parse_sconto(sc_ven_raw)

        if sc_acq is None or sc_ven is None:
            st.error("Inserire valori numerici validi (es. 45 oppure 40+10+5).")
        elif (1 - sc_ven) == 0:
            st.error("Sconto vendita del 100%: impossibile calcolare.")
        else:
            margine2  = (sc_acq - sc_ven) / (1 - sc_ven)           # MARGINE
            ricarico2 = ((1 - sc_ven) / (1 - sc_acq)) - 1 if sc_acq < 1 else 0  # RICARICO

            ca, cb = st.columns(2)
            with ca:
                cls = "ok-val" if margine2 >= float(T["soglia_margine"]) / 100 else "danger-v"
                result_box(str(T["tab2_res1"]), fmt_pct(margine2), cls)
            with cb:
                result_box(str(T["tab2_res2"]), fmt_pct(ricarico2), "secondary")

            if listino2 and listino2 > 0:
                acq_netto = listino2 * (1 - sc_acq)
                ven_netto = listino2 * (1 - sc_ven)
                cc, cd = st.columns(2)
                with cc:
                    result_box(str(T["tab2_res3"]), fmt_eur(acq_netto), "secondary")
                with cd:
                    result_box(str(T["tab2_res4"]), fmt_eur(ven_netto), "secondary")

            alert_margine_da_sconti(sc_acq, sc_ven)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — Prezzo di Vendita da Margine Desiderato
# Formula: Ven = Acq / (1 - Margine)
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown(
        f'<div class="section-label">Sezione {T["tab3_label_num"]}</div>'
        f'<div class="section-title">{T["tab3_titolo"]}</div>',
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    with c1:
        acq3 = st.number_input(
            str(T["tab3_input1"]), min_value=0.0, step=0.01, format="%.2f",
            key="s3_acq", value=None, placeholder="es. 50.00",
        )
    with c2:
        marg3 = st.number_input(
            str(T["tab3_input2"]), min_value=0.0, max_value=99.9, step=0.1, format="%.1f",
            key="s3_marg", value=None, placeholder="es. 15",
        )

    if acq3 is not None and marg3 is not None:
        if acq3 == 0:
            st.error("Il prezzo di acquisto non può essere zero.")
        elif marg3 >= 100:
            st.error("Il margine non può essere pari o superiore al 100%.")
        else:
            m = marg3 / 100
            ven3     = acq3 / (1 - m)          # PREZZO VENDITA
            ricarico3 = m / (1 - m)             # RICARICO equivalente
            utile3   = ven3 - acq3

            ca, cb, cc = st.columns(3)
            with ca:
                result_box(str(T["tab3_res1"]), fmt_eur(ven3), "ok-val")
            with cb:
                result_box(str(T["tab3_res2"]), fmt_pct(ricarico3), "secondary")
            with cc:
                result_box(str(T["tab3_res3"]), fmt_eur(utile3), "secondary")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — Calcolo Sconto da Listino e Netto
# Formula: Sconto = 1 - Netto/Listino
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    st.markdown(
        f'<div class="section-label">Sezione {T["tab4_label_num"]}</div>'
        f'<div class="section-title">{T["tab4_titolo"]}</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        f"<small style='color:{C['muted']}'>{T['tab4_hint']}</small><br><br>",
        unsafe_allow_html=True,
    )

    c1, c2 = st.columns(2)
    with c1:
        listino4 = st.number_input(
            str(T["tab4_input1"]), min_value=0.0, step=0.01, format="%.2f",
            key="s4_listino", value=None, placeholder="es. 100.00",
        )
    with c2:
        netto4 = st.number_input(
            str(T["tab4_input2"]), min_value=0.0, step=0.01, format="%.2f",
            key="s4_netto", value=None, placeholder="es. 60.00",
        )

    if listino4 is not None and netto4 is not None:
        if listino4 == 0:
            st.error("Il prezzo listino non può essere zero.")
        elif netto4 > listino4:
            st.warning("⚠ Il prezzo netto è maggiore del listino.")
        else:
            sconto4   = 1 - netto4 / listino4      # SCONTO SUL LISTINO
            risparmio = listino4 - netto4

            ca, cb = st.columns(2)
            with ca:
                result_box(str(T["tab4_res1"]), fmt_pct(sconto4), "ok-val")
            with cb:
                result_box(str(T["tab4_res2"]), fmt_eur(risparmio), "secondary")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — Sconti Concatenati (Cascata)
# Formula: netto = 100*(1-s1)*(1-s2)*(1-s3)...  sconto_tot = 1-netto/100
# ══════════════════════════════════════════════════════════════════════════════
with tab5:
    st.markdown(
        f'<div class="section-label">Sezione {T["tab5_label_num"]}</div>'
        f'<div class="section-title">{T["tab5_titolo"]}</div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        f"<small style='color:{C['muted']}'>{T['tab5_hint']}</small><br><br>",
        unsafe_allow_html=True,
    )

    sc5_raw = st.text_input(
        str(T["tab5_input1"]), key="s5_sc", placeholder="es.  40+10+5",
    )

    if sc5_raw and sc5_raw.strip():
        parti = [p.strip() for p in sc5_raw.split("+") if p.strip()]
        try:
            valori = [float(p) for p in parti]
            if any(v < 0 or v >= 100 for v in valori):
                st.error("Ogni sconto deve essere tra 0 e 99.")
            else:
                # Calcolo a cascata
                netto = 100.0
                for v in valori:
                    netto *= (1 - v / 100)
                sconto_tot = 1 - netto / 100

                ca, cb, cc = st.columns(3)
                with ca:
                    result_box(str(T["tab5_res1"]), fmt_pct(sconto_tot), "ok-val")
                with cb:
                    result_box(str(T["tab5_res2"]), f"{netto:.4f}", "secondary")
                with cc:
                    result_box(str(T["tab5_res3"]), str(len(valori)), "secondary")

                # Tabella dettaglio passo-passo
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown(
                    f"<div style='font-family:{F['font_mono']},monospace;"
                    f"font-size:{F['size_mono_small']};color:{C['muted']};'>"
                    f"{T['tab5_dettaglio']}</div>",
                    unsafe_allow_html=True,
                )
                base = 100.0
                righe = ""
                for i, v in enumerate(valori):
                    dopo  = base * (1 - v / 100)
                    righe += (
                        f"<tr>"
                        f"<td style='padding:0.4rem 1rem 0.4rem 0;color:{C['muted']}'>"
                        f"Sconto {i+1}</td>"
                        f"<td style='padding:0.4rem 1rem;color:{C['accent']}'>- {v:.1f}%</td>"
                        f"<td style='padding:0.4rem 1rem;color:{C['text']}'>→ {dopo:.4f}</td>"
                        f"</tr>"
                    )
                    base = dopo
                st.markdown(
                    f"<table style='border-collapse:collapse;"
                    f"font-family:{F['font_mono']},monospace;font-size:0.8rem;"
                    f"margin-top:0.5rem'>{righe}</table>",
                    unsafe_allow_html=True,
                )

        except ValueError:
            st.error("Formato non valido. Usa numeri separati da '+', es. 40+10+5")


# ══════════════════════════════════════════════════════════════════════════════
# FOOTER — testi da configurazione
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("<br><br>", unsafe_allow_html=True)
st.markdown(
    f"<div style='border-top:1px solid {C['border']};padding-top:1rem;"
    f"display:flex;justify-content:space-between;align-items:center'>"
    f"<span style='font-family:{F['font_mono']},monospace;"
    f"font-size:{F['size_mono_small']};color:{C['muted']};letter-spacing:0.1em'>"
    f"{T['footer_sx']}</span>"
    f"<span style='font-family:{F['font_mono']},monospace;"
    f"font-size:{F['size_mono_small']};color:{C['muted']}'>"
    f"{T['footer_dx']}</span>"
    f"</div>",
    unsafe_allow_html=True,
)
