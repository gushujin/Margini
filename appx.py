"""
Calcolatrice Commerciale — app.py
Versione Aggiornata: Focus su Risorsa da richiedere
"""

import streamlit as st
import openpyxl

st.set_page_config(
    page_title="Calcolatrice Commerciale",
    page_icon="🧮",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ══════════════════════════════════════════════════════════════════════════════
# CARICAMENTO CONFIGURAZIONE DA config.xlsx
# ══════════════════════════════════════════════════════════════════════════════

@st.cache_data
def load_config():
    wb = openpyxl.load_workbook("config.xlsx", data_only=True)

    colori = {}
    for row in wb["COLORI"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            colori[str(row[0]).strip()] = str(row[1]).strip()

    font = {}
    for row in wb["FONT"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            font[str(row[0]).strip()] = str(row[1]).strip()

    testi = {}
    for row in wb["TESTI"].iter_rows(min_row=2, values_only=True):
        if row[0]:
            testi[str(row[0]).strip()] = row[1]

    return colori, font, testi


C, F, T = load_config()

# ══════════════════════════════════════════════════════════════════════════════
# CSS GENERATO DINAMICAMENTE
# ══════════════════════════════════════════════════════════════════════════════

weights_display = F["font_display_weights"].replace(";", ";")
weights_mono    = F["font_mono_weights"].replace(";", ";")
google_url = (
    f"https://fonts.googleapis.com/css2?"
    f"family={F['font_display'].replace(' ', '+')}:wght@{weights_display}"
    f"&family={F['font_mono'].replace(' ', '+')}:wght@{weights_mono}"
    f"&display=swap"
)

css = f"""
<style>
@import url('{google_url}');

:root {{
    --bg:        {C['bg']};
    --surface:   {C['surface']};
    --border:    {C['border']};
    --accent:    {C['accent']};
    --accent2:   {C['accent2']};
    --text:      {C['text']};
    --muted:     {C['muted']};
    --danger:    {C['danger']};
    --ok:        {C['ok']};
    --f-display: '{F['font_display']}', sans-serif;
    --f-mono:    '{F['font_mono']}', monospace;
    --sz-title:      {F['size_title']};
    --sz-section:    {F['size_section']};
    --sz-result:     {F['size_result']};
    --sz-result-sec: {F['size_result_sec']};
    --sz-label:      {F['size_label']};
    --sz-mono-sm:    {F['size_mono_small']};
}}

html, body, [data-testid="stAppViewContainer"] {{
    background-color: var(--bg) !important;
    color: var(--text);
    font-family: var(--f-display);
}}

[data-testid="stHeader"]  {{ background: transparent !important; }}
[data-testid="stSidebar"] {{ background: var(--surface) !important; }}
#MainMenu, footer, header {{ visibility: hidden; }}

.topbar {{
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 1.5rem 0 2rem;
    border-bottom: 1px solid var(--border);
    margin-bottom: 2rem;
}}
.topbar h1 {{
    font-family: var(--f-display);
    font-size: var(--sz-title);
    font-weight: 800;
    margin: 0;
    color: var(--text);
}}
.topbar .tag {{
    background: var(--accent);
    color: #000;
    font-family: var(--f-mono);
    font-size: var(--sz-mono-sm);
    padding: 0.3rem 0.7rem;
    border-radius: 2px;
}}

.section-label {{
    font-family: var(--f-mono);
    font-size: var(--sz-mono-sm);
    letter-spacing: 0.18em;
    text-transform: uppercase;
    color: var(--accent);
    margin-bottom: 0.4rem;
}}
.section-title {{
    font-size: var(--sz-section);
    font-weight: 700;
    margin-bottom: 1.25rem;
    color: var(--text);
}}

.result-box {{
    background: #111;
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 1rem 1.2rem;
    margin-top: 0.5rem;
}}
.result-label {{
    font-family: var(--f-mono);
    font-size: var(--sz-mono-sm);
    color: var(--muted);
    margin-bottom: 0.25rem;
    text-transform: uppercase;
}}
.result-value {{ font-size: var(--sz-result); font-weight: 800; color: var(--accent); font-family: var(--f-display); }}
.result-value.secondary {{ font-size: var(--sz-result-sec); color: var(--text); }}
.result-value.ok-val {{ color: var(--ok); }}
.result-value.danger-v {{ color: var(--danger); }}

.alert-box {{
    background: rgba(255,59,59,0.08);
    border: 1px solid rgba(255,59,59,0.35);
    border-left: 3px solid var(--danger);
    border-radius: 6px;
    padding: 1.2rem;
    margin-top: 1rem;
}}

/* Stile base dei campi di input */
div[data-testid="stNumberInput"] input,
div[data-testid="stTextInput"] input {{
    background: #111 !important;
    border: 1px solid var(--border) !important;
    color: var(--text) !important;
    font-family: var(--f-mono) !important;
    transition: all 0.2s ease-in-out !important;
}}

/* EVIDENZIAZIONE: Quando clicchi o scrivi nel campo */
div[data-testid="stNumberInput"]:focus-within, 
div[data-testid="stTextInput"]:focus-within {{
    border: 1px solid var(--accent) !important;
    box-shadow: 0 0 12px rgba(255, 255, 255, 0.15) !important;
}}

/* Cambia colore alla Label quando il campo è attivo */
div[data-testid="stNumberInput"]:focus-within label, 
div[data-testid="stTextInput"]:focus-within label {{
    color: var(--accent) !important;
}}


label[data-testid="stWidgetLabel"] {{
    color: var(--muted) !important;
    font-family: var(--f-mono) !important;
    font-size: var(--sz-label) !important;
    text-transform: uppercase !important;
}}

button[data-baseweb="tab"] {{
    font-family: var(--f-display) !important;
    font-weight: 600 !important;
}}

div[data-testid="stButton"] > button {{
    background: transparent !important;
    border: 1px solid var(--border) !important;
    color: var(--muted) !important;
    font-family: var(--f-mono) !important;
}}
</style>
"""
st.markdown(css, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def parse_sconto(s: str):
    if not s or not s.strip(): return None
    try:
        netto = 100.0
        for parte in s.strip().split("+"):
            v = float(parte.strip())
            if v < 0 or v >= 100: return None
            netto *= (1 - v / 100)
        return 1 - netto / 100
    except: return None

def fmt_pct(v: float) -> str: return f"{v * 100:.2f}%"
def fmt_eur(v: float) -> str: return f"€ {v:,.2f}"

def result_box(label: str, value: str, cls: str = ""):
    st.markdown(f'<div class="result-box"><div class="result-label">{label}</div><div class="result-value {cls}">{value}</div></div>', unsafe_allow_html=True)

def alert_margine(acquisto: float, vendita: float):
    soglia = float(T["soglia_margine"]) / 100
    if vendita <= 0: return
    margine = (vendita - acquisto) / vendita
    if margine < soglia:
        # Calcolo del prezzo d'acquisto necessario per arrivare alla soglia
        acq_max = vendita * (1 - soglia)
        # Risorsa necessaria (differenza tra acquisto attuale e acquisto necessario)
        risorsa = acquisto - acq_max
        
        st.markdown(
            f'<div class="alert-box">'
            f'<div style="color:var(--danger); font-weight:700; font-family:var(--f-display); margin-bottom:0.5rem; font-size:1rem;">{T["alert_title"]}</div>'
            f'<div style="font-family:var(--f-mono); font-size:0.9rem; line-height:1.6;">'
            f'Margine attuale: <span style="color:var(--danger); font-weight:700;">{margine*100:.2f}%</span> (Minimo: {T["soglia_margine"]}%)<br>'
            f'Richiedere risorsa al fornitore: <span style="color:var(--text); font-weight:800; font-size:1.1rem;">{fmt_eur(risorsa)}</span><br>'
            f'<span style="color:var(--muted); font-size:0.8rem;">(Massimo prezzo d\'acquisto consentito: {fmt_eur(acq_max)})</span>'
            f'</div></div>', 
            unsafe_allow_html=True
        )

# ══════════════════════════════════════════════════════════════════════════════
# TOP BAR & STATE
# ══════════════════════════════════════════════════════════════════════════════

col_h, col_btn = st.columns([5, 1])
with col_h:
    st.markdown(f'<div class="topbar"><h1>{T["app_title"]}</h1><span class="tag">{T["app_version"]}</span></div>', unsafe_allow_html=True)
with col_btn:
    st.markdown("<div style='padding-top:1.9rem'>", unsafe_allow_html=True)
    if st.button(str(T["btn_reset"])):
        for k in st.session_state.keys(): del st.session_state[k]
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# LAYOUT A 2 TABS
# ══════════════════════════════════════════════════════════════════════════════

tab_main, tab_multi = st.tabs(["📊 Calcoli Rapidi", str(T["tab5_nome"])])

with tab_main:
    # --- SEZIONE 1: MARGINE DA PREZZI ---
    st.markdown(f'<div class="section-label">Sezione {T["tab1_label_num"]}</div><div class="section-title">{T["tab1_titolo"]}</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1: acq1 = st.number_input(str(T["tab1_input1"]), min_value=0.0, step=0.01, key="s1_acq", value=None)
    with c2: ven1 = st.number_input(str(T["tab1_input2"]), min_value=0.0, step=0.01, key="s1_ven", value=None)
    if acq1 and ven1:
        margine = (ven1 - acq1) / ven1
        ca, cb, cc = st.columns(3)
        with ca: result_box(str(T["tab1_res1"]), fmt_pct(margine), "ok-val" if margine >= float(T["soglia_margine"])/100 else "danger-v")
        with cb: result_box(str(T["tab1_res2"]), fmt_pct((ven1/acq1-1)) if acq1 > 0 else "0%", "secondary")
        with cc: result_box(str(T["tab1_res3"]), fmt_eur(ven1-acq1), "secondary")
        alert_margine(acq1, ven1)

    st.markdown("<hr style='border: 0; border-top: 1px solid var(--border); margin: 2rem 0;'>", unsafe_allow_html=True)

    # --- SEZIONE 2: MARGINE DA SCONTI ---
    st.markdown(f'<div class="section-label">Sezione {T["tab2_label_num"]}</div><div class="section-title">{T["tab2_titolo"]}</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1: sc_acq_raw = st.text_input(str(T["tab2_input1"]), key="s2_sc_acq")
    with c2: sc_ven_raw = st.text_input(str(T["tab2_input2"]), key="s2_sc_ven")
    if sc_acq_raw and sc_ven_raw:
        sa, sv = parse_sconto(sc_acq_raw), parse_sconto(sc_ven_raw)
        if sa is not None and sv is not None:
            marg2 = (sa - sv) / (1 - sv) if (1-sv) != 0 else 0
            ca, cb = st.columns(2)
            with ca: result_box(str(T["tab2_res1"]), fmt_pct(marg2), "ok-val" if marg2 >= float(T["soglia_margine"])/100 else "danger-v")
            with cb: result_box(str(T["tab2_res2"]), fmt_pct(((1-sv)/(1-sa)-1)) if (1-sa) != 0 else "0%", "secondary")

    st.markdown("<hr style='border: 0; border-top: 1px solid var(--border); margin: 2rem 0;'>", unsafe_allow_html=True)

    # --- SEZIONE 3: PREZZO DA MARGINE ---
    st.markdown(f'<div class="section-label">Sezione {T["tab3_label_num"]}</div><div class="section-title">{T["tab3_titolo"]}</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1: acq3 = st.number_input(str(T["tab3_input1"]), min_value=0.0, key="s3_acq", value=None)
    with c2: marg3 = st.number_input(str(T["tab3_input2"]), min_value=0.0, key="s3_marg", value=None)
    if acq3 and marg3:
        v3 = acq3 / (1 - marg3/100) if marg3 < 100 else 0
        ca, cb = st.columns(2)
        with ca: result_box(str(T["tab3_res1"]), fmt_eur(v3), "ok-val")
        with cb: result_box("Utile", fmt_eur(v3-acq3), "secondary")

    st.markdown("<hr style='border: 0; border-top: 1px solid var(--border); margin: 2rem 0;'>", unsafe_allow_html=True)

    # --- SEZIONE 4: CALCOLO SCONTO ---
    st.markdown(f'<div class="section-label">Sezione {T["tab4_label_num"]}</div><div class="section-title">{T["tab4_titolo"]}</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1: li4 = st.number_input(str(T["tab4_input1"]), min_value=0.0, key="s4_li", value=None)
    with c2: ne4 = st.number_input(str(T["tab4_input2"]), min_value=0.0, key="s4_ne", value=None)
    if li4 and ne4:
        result_box(str(T["tab4_res1"]), fmt_pct(1 - ne4/li4) if li4 > 0 else "0%", "ok-val")

# ── TAB SCONTI MULTIPLI ──
with tab_multi:
    st.markdown(f'<div class="section-label">Sezione 5</div><div class="section-title">{T["tab5_titolo"]}</div>', unsafe_allow_html=True)
    sc5_raw = st.text_input(str(T["tab5_input1"]), key="s5_sc")
    if sc5_raw:
        s_tot = parse_sconto(sc5_raw)
        if s_tot: result_box(str(T["tab5_res1"]), fmt_pct(s_tot), "ok-val")

# ══════════════════════════════════════════════════════════════════════════════
# FOOTER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(f"<br><div style='border-top:1px solid {C['border']};padding-top:1rem;color:{C['muted']};font-size:0.8rem'>{T['footer_sx']}</div>", unsafe_allow_html=True)
